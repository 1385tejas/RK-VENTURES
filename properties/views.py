from django.shortcuts import render, get_object_or_404, redirect
from .models import Property
import openpyxl
from django.core.files.storage import FileSystemStorage
from django.conf import settings
import os
from django.http import HttpResponseForbidden
from notifications.models import Notification
from django.contrib.auth.models import User
from django.core.mail import send_mail

def create_notification_for_all_users(title, message, property_obj=None, send_email=False):
    for user in User.objects.all():
        Notification.objects.create(user=user, title=title, message=message, property=property_obj)
        if send_email and user.email:
            try:
                property_link = f"{settings.SITE_DOMAIN}/properties/{property_obj.id}/" if property_obj else ''
                html_message = f"""
                <p>Hello {user.get_full_name() or user.username},</p>
                <p>{message}</p>
                {f'<p><a href="{property_link}" style="color:#8a1832;font-weight:bold;">Click here to view the property and auction details.</a></p>' if property_link else ''}
                <br><p>Best regards,<br>RK Ventures Team</p>
                """
                send_mail(
                    subject=title,
                    message=message + (f"\nView property: {property_link}" if property_link else ''),
                    from_email=None,
                    recipient_list=[user.email],
                    html_message=html_message,
                    fail_silently=False,  # Show errors for debugging
                )
            except Exception:
                pass

def normalize_auction_date(properties):
    for prop in properties:
        ed = prop.extra_data
        if 'auction_date' not in ed:
            if 'Auction_date' in ed:
                ed['auction_date'] = ed['Auction_date']
            elif 'auction date' in ed:
                ed['auction_date'] = ed['auction date']

def home(request):
    user = request.user
    properties = Property.objects.all().order_by('-created_at')
    normalize_auction_date(properties)
    context = {
        'properties': properties,
        'user': user,
    }
    return render(request, 'properties/home.html', context)

def about(request):
    return render(request, 'properties/about.html')

def contact(request):
    return render(request, 'properties/contact.html')

def properties_list(request):
    properties = Property.objects.all().order_by('-created_at')
    normalize_auction_date(properties)
    context = {
        'properties': properties,
        'user': request.user,
    }
    return render(request, 'properties/properties_list.html', context)

def property_detail(request, property_id):
    property_obj = get_object_or_404(Property, id=property_id)
    context = {
        'property': property_obj,
        'user': request.user,
    }
    return render(request, 'properties/property_detail.html', context)

def upload_properties(request):
    if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden('You do not have permission to upload properties.')
    message = None
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        fs = FileSystemStorage()
        filename = fs.save(excel_file.name, excel_file)
        file_path = fs.path(filename)
        added_count = 0
        added_ids = []
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                values = list(row)
                if not any(values):
                    continue  # skip empty rows
                prop_data = {headers[j]: (values[j] if j < len(values) and values[j] is not None else '') for j in range(len(headers))}
                # Convert all extra_data values to strings
                for k in list(prop_data.keys()):
                    v = prop_data[k]
                    try:
                        from datetime import datetime
                        if isinstance(v, datetime):
                            prop_data[k] = v.strftime('%Y-%m-%d')
                        else:
                            prop_data[k] = str(v)
                    except Exception:
                        prop_data[k] = str(v)
                # Extract core fields
                prop_id = prop_data.pop('ID', '') or prop_data.pop('id', '')
                title = prop_data.pop('Title', '') or prop_data.pop('title', '')
                price = prop_data.pop('Price', '') or prop_data.pop('price', '')
                location = prop_data.pop('Location', '') or prop_data.pop('location', '')
                description = prop_data.pop('Description', '') or prop_data.pop('description', '')
                image_url = prop_data.pop('Image', '') or prop_data.pop('image', '')
                extra_images = prop_data.pop('Images', [])
                if isinstance(extra_images, str):
                    extra_images = [extra_images]
                if image_url:
                    extra_images = [image_url] + extra_images
                # Always add property (no duplicate check)
                prop = Property.objects.create(
                    title=title.strip() if title else '',
                    price=price,
                    location=location.strip() if location else '',
                    description=description,
                    extra_images=extra_images,
                    extra_data=prop_data
                )
                added_count += 1
                added_ids.append(str(prop.id))
                # Send notification for each property added
                auction_date = prop.extra_data.get('auction_date') or prop.extra_data.get('Auction_date') or prop.extra_data.get('auction date')
                msg = f"New property added: {prop.title or 'Untitled'}"
                if auction_date:
                    # Format auction_date as string if it's a datetime object
                    try:
                        from datetime import datetime
                        if isinstance(auction_date, datetime):
                            auction_date_str = auction_date.strftime('%Y-%m-%d')
                        else:
                            auction_date_str = str(auction_date)
                    except Exception:
                        auction_date_str = str(auction_date)
                    msg += f" | Auction Date: {auction_date_str}"
                create_notification_for_all_users('New Property Added', msg, prop, send_email=True)
            if added_count:
                message = f"Properties added: {added_count} (IDs: {', '.join(added_ids)})."
            else:
                message = f"No properties added."
        except Exception as e:
            import traceback
            print('Error processing file:', traceback.format_exc())
            message = 'Error processing file: Please check your data format.'
        fs.delete(filename)
    return render(request, 'properties/upload_properties.html', {'message': message})

def property_edit(request, property_id):
    if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden('You do not have permission to edit properties.')
    property_obj = get_object_or_404(Property, id=property_id)
    message = None
    if request.method == 'POST':
        # Update core fields
        property_obj.title = request.POST.get('title', property_obj.title)
        property_obj.price = request.POST.get('price', property_obj.price)
        property_obj.location = request.POST.get('location', property_obj.location)
        property_obj.description = request.POST.get('description', property_obj.description)
        # Update extra_data
        for key in request.POST.keys():
            if key not in ['id', 'csrfmiddlewaretoken', 'title', 'price', 'location', 'description']:
                property_obj.extra_data[key] = request.POST.get(key)
        # Handle multiple image uploads
        images = property_obj.extra_images or []
        uploaded_files = request.FILES.getlist('Images')
        for img_file in uploaded_files:
            fs = FileSystemStorage()
            filename = fs.save(img_file.name, img_file)
            images.append(fs.url(filename))
        if uploaded_files:
            property_obj.extra_images = images
            if not property_obj.main_image:
                property_obj.main_image = uploaded_files[0]
        # Handle main image upload
        if request.FILES.get('Image'):
            img_file = request.FILES['Image']
            fs = FileSystemStorage()
            filename = fs.save(img_file.name, img_file)
            property_obj.main_image = filename
            images.insert(0, fs.url(filename))
            property_obj.extra_images = images
        property_obj.save()
        message = 'Property updated successfully!'
    context = {
        'property': property_obj,
        'user': request.user,
        'message': message,
    }
    return render(request, 'properties/property_edit.html', context)

def property_delete(request, property_id):
    if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden('You do not have permission to delete properties.')
    property_obj = get_object_or_404(Property, id=property_id)
    property_obj.delete()
    return render(request, 'properties/property_deleted.html', {'property_id': property_id})
